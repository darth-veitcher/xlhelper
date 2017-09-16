"""Efficient interaction with Excel files for ETL purposes.

Aim is to process without using pandas and other 'heavy' library requirements
whilst adhering to best practices.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from collections import OrderedDict


def sql_safe_string(string):
    """Process a string and strip non-safe chars from it for SQL."""
    safe_chars = [32, ]  # space
    [safe_chars.append(i) for i in range(48, 58)]  # 0 to 9
    [safe_chars.append(i) for i in range(65, 91)]  # A to Z
    [safe_chars.append(i) for i in range(97, 122)]  # a to z

    if not string:
        return ""

    output = ""
    for c in string:
        if ord(c) in safe_chars:
            output = output + c.replace(" ", "_")
    return output


def sheet_to_dict(file_path, sheet_name=None, header_row=1, start_col='A',
                  sql_safe=False, keep_order=False, remapping_dict=None):
    """Import a sheet from specified Excel file and return a generator for a
    list of Python dictionaries with k,v corresponding to header and row vals.

    Params:
        - file_path:              absolute path to the file
        - sheet_name (optional):  the sheet to import (default is the first)
        - header_row (optional):  the position of the header row (default is 1)
        - start_col (optional):   the first column of data (default is A)
        - sql_safe (optional):    will convert keys to a sql_safe string,
                                  replacing spaces with underscores
                                  (default is False)
        - keep_order (optional):  uses an OrderedDict to keep the data
                                  structure (column ordering) consistent with
                                  the spreadsheet (default is True)
    """
    # Optimised loader, use the data_only flag to ensure we get the values,
    # otherwise it will return the formula string in the cell as opposed to
    # what that formula resolves to...
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        ws = wb.get_sheet_by_name(sheet_name)
    else:
        ws = wb.active  # default to the first sheet in file

    # Create list of header cell values to use as keys for the dict
    # '_max_column': 16,
    # '_max_row': 162,
    # '_min_column': 1,
    # '_min_row': 1
    try:
        header_range = "{start_col}{header_row}:{col}{header_row}".format(
            start_col=start_col,
            header_row=header_row,
            col=get_column_letter(ws._max_column)
        )
    except TypeError as e:
        print(e)
        print("Did you set the Header row correctly?!")

    if sql_safe:
        headers = [sql_safe_string(cell.value) for row in
                   ws[header_range] for cell in row]
    else:
        headers = [cell.value for row in ws[header_range]
                   for cell in row]

    # Remap columns
    if remapping_dict:
        i = 0
        for h in headers:
            if h in remapping_dict:
                headers[i] = remapping_dict[h]
            i = i + 1
    headers = tuple(headers)

    # Get the rows
    # '_max_column': 16,
    # '_max_row': 162,
    # '_min_column': 1,
    # '_min_row': 1
    rows_range = {
        'start_col': start_col,
        'end_col': get_column_letter(ws._max_column),
        'start_row': header_row + 1,
        'end_row': ws._max_row + 1
    }

    # Return generator
    # Create dicts for the rows using header information as the keys
    for index, row in enumerate(range(rows_range['start_row'],
                                      rows_range['end_row'])):
        row_range = '{start_col}{row}:{end_col}{row}'.format(
            start_col=rows_range['start_col'],
            end_col=rows_range['end_col'],
            row=row
        )
        r = ws[row_range][0]

        if keep_order:
            yield OrderedDict(zip(headers, [cell.value for cell in r]))
        else:
            yield dict(zip(headers, [cell.value for cell in r]))
